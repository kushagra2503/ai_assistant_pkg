import os
import logging
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from typing import Optional, List, Dict, Any, Tuple, Union
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
import xlwings as xw

logger = logging.getLogger("ai_assistant")

class ExcelHandler:
    """A class for handling Excel file operations."""
    
    def __init__(self, workspace_root: Optional[str] = None):
        """
        Initialize the Excel handler.
        
        Args:
            workspace_root: The root directory of the workspace.
        """
        self.workspace_root = workspace_root or os.getcwd()
        
    def resolve_path(self, file_path: str) -> str:
        """Resolve a relative path to an absolute path."""
        if os.path.isabs(file_path):
            return file_path
        return os.path.join(self.workspace_root, file_path)
    
    def list_excel_files(self, directory: str = "") -> List[str]:
        """
        List all Excel files in the specified directory.
        
        Args:
            directory: Directory to search in (relative to workspace root).
            
        Returns:
            List of Excel file paths.
        """
        try:
            dir_path = self.resolve_path(directory)
            excel_files = []
            
            for root, _, files in os.walk(dir_path):
                for file in files:
                    if file.endswith(('.xlsx', '.xls', '.xlsm')):
                        rel_path = os.path.relpath(os.path.join(root, file), self.workspace_root)
                        excel_files.append(rel_path)
                        
            return excel_files
        except Exception as e:
            logger.error(f"Error listing Excel files: {e}")
            return []
    
    def read_excel_file(self, file_path: str, sheet_name: Optional[Union[str, int]] = 0) -> Optional[pd.DataFrame]:
        """
        Read an Excel file and return its contents as a DataFrame.
        
        Args:
            file_path: Path to the Excel file.
            sheet_name: Name or index of the sheet to read.
            
        Returns:
            DataFrame containing the Excel data or None if there's an error.
        """
        try:
            abs_path = self.resolve_path(file_path)
            return pd.read_excel(abs_path, sheet_name=sheet_name)
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {e}")
            return None
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Get all sheet names from an Excel file.
        
        Args:
            file_path: Path to the Excel file.
            
        Returns:
            List of sheet names.
        """
        try:
            abs_path = self.resolve_path(file_path)
            xls = pd.ExcelFile(abs_path)
            return xls.sheet_names
        except Exception as e:
            logger.error(f"Error getting sheet names from {file_path}: {e}")
            return []
    
    def get_excel_info(self, file_path: str) -> Dict[str, Any]:
        """
        Get information about an Excel file.
        
        Args:
            file_path: Path to the Excel file.
            
        Returns:
            Dictionary containing file information.
        """
        try:
            abs_path = self.resolve_path(file_path)
            file_info = {
                "file_name": os.path.basename(abs_path),
                "file_size": os.path.getsize(abs_path),
                "sheets": [],
                "last_modified": os.path.getmtime(abs_path)
            }
            
            # Get sheet information
            workbook = openpyxl.load_workbook(abs_path, read_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info = {
                    "name": sheet_name,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column
                }
                file_info["sheets"].append(sheet_info)
                
            return file_info
        except Exception as e:
            logger.error(f"Error getting Excel info for {file_path}: {e}")
            return {"error": str(e)}
    
    def save_dataframe_to_excel(self, df: pd.DataFrame, file_path: str, sheet_name: str = "Sheet1") -> bool:
        """
        Save a DataFrame to an Excel file.
        
        Args:
            df: DataFrame to save.
            file_path: Path to save the Excel file.
            sheet_name: Name of the sheet.
            
        Returns:
            True if successful, False otherwise.
        """
        try:
            abs_path = self.resolve_path(file_path)
            df.to_excel(abs_path, sheet_name=sheet_name, index=False)
            logger.info(f"Successfully saved DataFrame to {file_path}")
            return True
        except Exception as e:
            logger.error(f"Error saving DataFrame to Excel file {file_path}: {e}")
            return False
    
    def connect_to_excel_app(self, file_path: Optional[str] = None) -> Optional[xw.Book]:
        """
        Connect to the Excel application and open a file.
        
        Args:
            file_path: Path to the Excel file to open.
            
        Returns:
            xlwings Book object or None if there's an error.
        """
        try:
            if file_path:
                abs_path = self.resolve_path(file_path)
                return xw.Book(abs_path)
            else:
                return xw.Book()  # New workbook
        except Exception as e:
            logger.error(f"Error connecting to Excel app: {e}")
            return None

    def create_excel_chart(self, file_path: str, sheet_name: str, data_range: str, 
                         chart_type: str = "bar", title: str = "Chart") -> bool:
        """
        Create a chart in an Excel file.
        
        Args:
            file_path: Path to the Excel file.
            sheet_name: Name of the sheet containing the data.
            data_range: Range of cells containing the data (e.g., "A1:B10").
            chart_type: Type of chart ("bar", "line", "pie").
            title: Title of the chart.
            
        Returns:
            True if successful, False otherwise.
        """
        try:
            abs_path = self.resolve_path(file_path)
            wb = openpyxl.load_workbook(abs_path)
            
            if sheet_name not in wb.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in {file_path}")
                return False
                
            ws = wb[sheet_name]
            
            # Create chart based on type
            if chart_type.lower() == "bar":
                chart = BarChart()
            elif chart_type.lower() == "line":
                chart = LineChart()
            elif chart_type.lower() == "pie":
                chart = PieChart()
            else:
                logger.error(f"Unsupported chart type: {chart_type}")
                return False
                
            # Parse data range
            try:
                # Simple parsing for ranges like "A1:B10"
                data_parts = data_range.split(":")
                if len(data_parts) != 2:
                    raise ValueError(f"Invalid data range format: {data_range}")
                    
                from openpyxl.utils import column_index_from_string
                
                # Extract column letters and row numbers
                start_col_letter = ''.join(c for c in data_parts[0] if c.isalpha())
                start_row = int(''.join(c for c in data_parts[0] if c.isdigit()))
                
                end_col_letter = ''.join(c for c in data_parts[1] if c.isalpha())
                end_row = int(''.join(c for c in data_parts[1] if c.isdigit()))
                
                # Convert column letters to indices
                start_col = column_index_from_string(start_col_letter)
                end_col = column_index_from_string(end_col_letter)
                
                # Create data references
                data = Reference(ws, min_col=start_col, min_row=start_row,
                               max_col=end_col, max_row=end_row)
                
                # For category labels (first row)
                cats = Reference(ws, min_col=start_col, min_row=start_row,
                               max_col=end_col, max_row=start_row)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.title = title
                
                # Add chart to worksheet
                ws.add_chart(chart, "H1")  # Add chart at position H1
                
                wb.save(abs_path)
                logger.info(f"Successfully created {chart_type} chart in {file_path}")
                return True
                
            except ValueError as ve:
                logger.error(f"Error parsing data range: {ve}")
                return False
                
        except Exception as e:
            logger.error(f"Error creating chart in Excel file {file_path}: {e}")
            return False
    
    def run_excel_formula(self, file_path: str, sheet_name: str, 
                        cell: str, formula: str) -> Any:
        """
        Run a formula in an Excel file.
        
        Args:
            file_path: Path to the Excel file.
            sheet_name: Name of the sheet.
            cell: Cell to put the formula in (e.g., "A1").
            formula: Excel formula to run.
            
        Returns:
            Result of the formula.
        """
        try:
            # Connect to Excel application
            book = self.connect_to_excel_app(file_path)
            if not book:
                return None
                
            # Get the worksheet
            sheet = book.sheets[sheet_name]
            
            # Apply formula to cell
            sheet.range(cell).formula = formula
            
            # Get the result
            result = sheet.range(cell).value
            
            # Save and close
            book.save()
            book.close()
            
            return result
            
        except Exception as e:
            logger.error(f"Error running Excel formula: {e}")
            return None

    def analyze_excel_data(self, df: pd.DataFrame, analysis_type: str = "summary") -> Dict[str, Any]:
        """
        Analyze Excel data with various statistical methods.
        
        Args:
            df: DataFrame to analyze.
            analysis_type: Type of analysis to perform (summary, correlation, descriptive).
            
        Returns:
            Dictionary with analysis results.
        """
        try:
            result = {}
            
            if analysis_type == "summary":
                # Calculate summary statistics
                numeric_cols = df.select_dtypes(include=["number"])
                
                if numeric_cols.empty:
                    result["error"] = "No numeric columns available for analysis"
                    return result
                    
                # Generate summary statistics
                result["summary"] = numeric_cols.describe().to_dict()
                
                # Count null values
                result["null_values"] = df.isnull().sum().to_dict()
                
            elif analysis_type == "correlation":
                # Calculate correlation matrix
                numeric_cols = df.select_dtypes(include=["number"])
                
                if numeric_cols.empty:
                    result["error"] = "No numeric columns available for correlation analysis"
                    return result
                    
                result["correlation"] = numeric_cols.corr().to_dict()
                
            elif analysis_type == "descriptive":
                # Descriptive statistics for each column
                result["descriptive"] = {}
                
                for column in df.columns:
                    col_data = df[column]
                    col_stats = {}
                    
                    # Basic statistics
                    col_stats["count"] = len(col_data)
                    col_stats["null_count"] = col_data.isnull().sum()
                    col_stats["unique_values"] = col_data.nunique()
                    
                    # For numeric columns
                    if np.issubdtype(col_data.dtype, np.number):
                        col_stats["min"] = col_data.min()
                        col_stats["max"] = col_data.max()
                        col_stats["mean"] = col_data.mean()
                        col_stats["median"] = col_data.median()
                        col_stats["std_dev"] = col_data.std()
                        
                    # For string columns
                    elif col_data.dtype == object:
                        # Sample of unique values (up to 5)
                        unique_vals = col_data.dropna().unique()
                        col_stats["sample_values"] = list(unique_vals[:5])
                        col_stats["avg_length"] = col_data.astype(str).str.len().mean()
                        
                    result["descriptive"][column] = col_stats
                    
            else:
                result["error"] = f"Unsupported analysis type: {analysis_type}"
                
            return result
            
        except Exception as e:
            logger.error(f"Error analyzing Excel data: {e}")
            return {"error": str(e)}

    def query_excel_data(self, df: pd.DataFrame, query: str) -> pd.DataFrame:
        """
        Query a DataFrame using pandas query syntax.
        
        Args:
            df: DataFrame to query.
            query: Query string in pandas query syntax.
            
        Returns:
            Filtered DataFrame.
        """
        try:
            filtered_df = df.query(query)
            return filtered_df
        except Exception as e:
            logger.error(f"Error querying Excel data: {e}")
            return pd.DataFrame()

    def generate_excel_visualization(self, df: pd.DataFrame, chart_type: str, 
                                   x_column: str, y_column: str, 
                                   title: str = "Chart") -> str:
        """
        Generate a visualization from Excel data and save it as an image.
        
        Args:
            df: DataFrame containing the data.
            chart_type: Type of chart to create (bar, line, scatter, pie, etc.).
            x_column: Column to use for x-axis.
            y_column: Column to use for y-axis.
            title: Title of the chart.
            
        Returns:
            Path to the saved chart image or error message.
        """
        try:
            # Create figure and axes
            plt.figure(figsize=(10, 6))
            
            # Ensure x_column and y_column exist in the DataFrame
            if x_column not in df.columns:
                return f"Error: Column '{x_column}' not found in DataFrame"
            if y_column not in df.columns:
                return f"Error: Column '{y_column}' not found in DataFrame"
                
            # Make a copy to avoid modifying the original DataFrame
            plot_df = df.copy()
            
            # Convert y_column to numeric, handling errors
            plot_df[y_column] = pd.to_numeric(plot_df[y_column], errors='coerce')
            plot_df = plot_df.dropna(subset=[y_column])
            
            # Check if DataFrame is empty after cleaning
            if plot_df.empty:
                return f"Error: No valid data points remain after processing. Check if '{y_column}' contains valid numeric values."
            
            # Convert x_column to string to avoid comparison issues in pie charts
            plot_df[x_column] = plot_df[x_column].astype(str)
            
            # Generate the chart based on type
            if chart_type == "bar":
                # Group data for bar charts to handle duplicate x values
                if len(plot_df) > 1:
                    grouped_df = plot_df.groupby(x_column)[y_column].sum().reset_index()
                    grouped_df.plot(kind='bar', x=x_column, y=y_column, title=title, legend=True)
                else:
                    plot_df.plot(kind='bar', x=x_column, y=y_column, title=title, legend=True)
            elif chart_type == "line":
                # Sort data for line charts by x_column if it's date-like
                try:
                    if pd.api.types.is_datetime64_any_dtype(df[x_column]):
                        plot_df = plot_df.sort_values(by=x_column)
                except:
                    pass  # If sorting fails, continue with unsorted data
                plot_df.plot(kind='line', x=x_column, y=y_column, title=title, legend=True)
            elif chart_type == "scatter":
                plot_df.plot(kind='scatter', x=x_column, y=y_column, title=title, legend=True)
            elif chart_type == "pie":
                # For pie charts, we need to group by x_column and sum y_column
                pie_data = plot_df.groupby(x_column)[y_column].sum()
                if pie_data.empty:
                    return f"Error: No data to plot after grouping. Check if data contains valid groups."
                pie_data.plot(kind='pie', autopct='%1.1f%%', title=title)
            elif chart_type == "hist":
                if plot_df[y_column].dropna().empty:
                    return f"Error: No valid data points for histogram. Column '{y_column}' has no valid numeric values."
                plot_df[y_column].plot(kind='hist', title=title, legend=True)
            else:
                return f"Error: Unsupported chart type '{chart_type}'"
            
            # Add labels and title
            plt.xlabel(x_column)
            plt.ylabel(y_column)
            plt.title(title)
            plt.tight_layout()
            
            # Save the chart to a file
            import tempfile
            import os
            
            # Create temporary file
            temp_dir = tempfile.gettempdir()
            chart_file = os.path.join(temp_dir, f"excel_chart_{os.getpid()}.png")
            
            # Save the chart
            plt.savefig(chart_file)
            plt.close()
            
            logger.info(f"Chart saved to {chart_file}")
            return chart_file
            
        except Exception as e:
            logger.error(f"Error generating Excel visualization: {e}")
            return f"Error generating visualization: {str(e)}"
